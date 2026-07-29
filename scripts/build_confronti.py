#!/usr/bin/env python3
"""Genera data/peer_group.json e data/provenienza.json dai file ISTAT.

peer_group.json  — serie 2014-2024 dei comuni della stessa categoria turistica di
                   Cefalu, per confrontarla con i suoi pari invece che con la media
                   nazionale, che comprende Roma e Milano.
provenienza.json — arrivi e presenze 2024 della provincia di Palermo per paese estero
                   di residenza e per tipologia ricettiva.

Uso:  python3 scripts/build_confronti.py
"""
import json, os, sys
import openpyxl

BASE = os.path.join(os.path.dirname(__file__), "..")
XLSX_COM = os.path.join(BASE, "DCSC_Occupancy_in_collective_accommodation", "2. Dati comunali 2014-2024.xlsx")
XLSX_PRO = os.path.join(BASE, "DCSC_Occupancy_in_collective_accommodation", "4. Dati per provenienza 2024.xlsx")
CEFALU, PROVINCIA, ANNI = "082027", "PALERMO", list(range(2014, 2025))


def peer_group():
    ws = openpyxl.load_workbook(XLSX_COM, read_only=True, data_only=True)["Comuni Classificazione-Brand"]
    cat = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[6]:
            cat[str(r[6]).zfill(6)] = (r[7] or "").strip()
    mia = cat.get(CEFALU)
    if not mia:
        sys.exit("Categoria di Cefalu non trovata")

    # Panel bilanciato: solo i comuni con la serie completa, altrimenti l'indice
    # risentirebbe di comuni che entrano ed escono da un anno all'altro.
    inclusi, agg = [], {a: {"arr": 0, "pre": 0, "alb": 0, "ext": 0} for a in ANNI}
    for cod, c in cat.items():
        if c != mia:
            continue
        p = os.path.join(BASE, "data", "serie", cod + ".json")
        if not os.path.exists(p):
            continue
        d = json.load(open(p, encoding="utf-8"))
        val = lambda k, a: d[k][d["anni"].index(a)] if a in d["anni"] else None
        if not all(isinstance(val("pre_tot", a), (int, float)) and val("pre_tot", a) > 0 for a in ANNI):
            continue
        inclusi.append(cod)
        for a in ANNI:
            for k, src in (("arr", "arr_tot"), ("pre", "pre_tot"), ("alb", "pre_alb"), ("ext", "pre_ext")):
                v = val(src, a)
                if isinstance(v, (int, float)):
                    agg[a][k] += v

    return {
        "_fonte": "ISTAT — classificazione dei comuni per categoria turistica prevalente, "
                  "aggregata sulle serie comunali 2014-2024.",
        "_nota": "Solo i comuni con serie completa su tutti gli anni, cosi l'indice non "
                 "risente di comuni presenti a intermittenza.",
        "_rigenera": "python3 scripts/build_confronti.py",
        "categoria": mia,
        "comuni_categoria": sum(1 for c in cat.values() if c == mia),
        "comuni_inclusi": len(inclusi),
        "anni": ANNI,
        "arr": [round(agg[a]["arr"]) for a in ANNI],
        "pre": [round(agg[a]["pre"]) for a in ANNI],
        "alb": [round(agg[a]["alb"]) for a in ANNI],
        "ext": [round(agg[a]["ext"]) for a in ANNI],
    }


def provenienza():
    ws = openpyxl.load_workbook(XLSX_PRO, read_only=True, data_only=True)["Data"]
    righe = list(ws.iter_rows(min_row=1, values_only=True))
    paesi = [(i, str(c).strip()) for i, c in enumerate(righe[4]) if i >= 5 and c]
    per_paese, per_tipo = {}, {}
    for r in righe[5:]:
        if not r or not r[3] or PROVINCIA not in str(r[3]).upper():
            continue
        et = str(r[4] or "").strip()
        if not et:
            continue
        misura = "arr" if et.endswith("Arrivi") else ("pre" if et.endswith("Presenze") else None)
        if not misura:
            continue
        # "Arrivi" e "Presenze" da soli sono le righe di totale provinciale,
        # tutto il resto e "<tipologia> Arrivi" o "<tipologia> Presenze".
        tipo = "" if et in ("Arrivi", "Presenze") else et.rsplit(" ", 1)[0].strip()
        for i, nome in paesi:
            v = r[i]
            if not isinstance(v, (int, float)) or v <= 0:
                continue
            if tipo:                       # riga "Arrivi"/"Presenze" senza tipologia = totale
                per_tipo.setdefault(tipo, {"arr": 0, "pre": 0})[misura] += v
            else:
                per_paese.setdefault(nome, {"arr": 0, "pre": 0})[misura] += v

    # Il file mette sulla stessa colonna le regioni italiane e i paesi esteri; le prime
    # sono scritte tutte in maiuscolo. Vanno tenute separate, altrimenti le quote non
    # significano nulla.
    ordina = lambda d: dict(sorted(d.items(), key=lambda x: -x[1]["pre"]))
    regioni = ordina({k: v for k, v in per_paese.items() if k.isupper()})
    esteri  = ordina({k: v for k, v in per_paese.items() if not k.isupper()})
    tot_it = sum(v["pre"] for v in regioni.values())
    tot_es = sum(v["pre"] for v in esteri.values())

    return {
        "_fonte": "ISTAT — arrivi e presenze 2024 per luogo di residenza dei clienti, "
                  "dettaglio provinciale.",
        "_nota": "Il dato e provinciale: comprende tutta la provincia di Palermo, non il solo "
                 "comune di Cefalu. Serve a leggere la composizione dei mercati, non i volumi "
                 "del comune.",
        "_rigenera": "python3 scripts/build_confronti.py",
        "provincia": PROVINCIA.title(),
        "anno": 2024,
        "totale_italia": round(tot_it),
        "totale_estero": round(tot_es),
        "regioni_italiane": regioni,
        "paesi_esteri": esteri,
        "tipologie": ordina(per_tipo),
    }


def main():
    for nome, dati in (("peer_group", peer_group()), ("provenienza", provenienza())):
        dest = os.path.join(BASE, "data", nome + ".json")
        json.dump(dati, open(dest, "w", encoding="utf-8"), ensure_ascii=False, indent=1)
        print("scritto data/%s.json" % nome)
    print("  categoria:", json.load(open(os.path.join(BASE, "data", "peer_group.json"), encoding="utf-8"))["categoria"])
    pv = json.load(open(os.path.join(BASE, "data", "provenienza.json"), encoding="utf-8"))
    print("  paesi esteri:", len(pv["paesi_esteri"]), "| regioni italiane:",
          len(pv["regioni_italiane"]), "| tipologie:", len(pv["tipologie"]))


if __name__ == "__main__":
    main()
